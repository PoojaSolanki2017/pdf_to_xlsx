# Import libraries 
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader
from openpyxl import Workbook

# Set input and output file names 
input_file = 'AML.pdf'
output_file = 'pdf_to_excel.xlsx'


# Open PDF file
pdf_file = open(input_file,'rb')

# Read PDF file
input_pdf = PyPDF2.PdfFileReader(pdf_file)

# Declare main fields of PDF file
main_list = [
 'Disease Description:',
 'Specific Indication:',
 'Molecular Abnormality:',
 'Test:',
 'Chromosome:',
 'Gene Symbol:',
 'Test Detects:',
 'Methodology:',
 'NCCN Category of Evidence:',
 'Specimen Types:',
 'NCCN Recommendation - Clinical Decision:',
 'Test Purpose:',
 'When to Test:',
 'Guideline Page with Test Recommendation:',
 'Notes:',
 '!"#$%&#']


# Create a workbook 
wb = Workbook()

# Grab the active worksheet
ws = wb.active

# Write field headings into excel file
row_num=1
column_num=1
for i in range(len(main_list)-1):
        field = main_list[i]
        ws.cell(row=row_num, column=column_num, value=field[:-1])
        column_num += 1

# Count total page number of PDF file
total_pages = input_pdf.getNumPages()

# Extract data from PDF and Write it into excel file
row_num = 2
for i in range(total_pages):
        page = input_pdf.getPage(i)
        page_content = page.extractText()
        column_num = 1
        for i in range(len(main_list)-1):
                field = main_list[i]
                next_field = main_list[i+1]
                # Find position of fields from extracted text of PDF file
                field_pos = page_content.find(field)
                next_field_pos = page_content.find(next_field)
                # Find position of field values from extracted text of PDF file
                field_value_start_pos = field_pos+len(field)
                field_value_end_pos = next_field_pos
                # Extract field values
                field_value = page_content[field_value_start_pos:field_value_end_pos]
                # Write field values into Excel
                ws.cell(row = row_num, column = column_num, value = field_value)
                column_num += 1
        row_num += 1

pdf_file.close()

# Save excel file
wb.save(output_file)

